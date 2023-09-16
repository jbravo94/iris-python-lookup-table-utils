import iris
import sys
import openpyxl

global lookupTableGlobalName
lookupTableGlobalName = "^Ens.LookupTable"

def get_value(table, key):
    ref = iris.gref(lookupTableGlobalName)
    return ref[table, key]

def print_value(table, key):
    print(get_value(table, key))

def set_value(table, key, value):
    ref = iris.gref(lookupTableGlobalName)
    ref[table, key] = value

def delete_key(table_name, key):
    ref = iris.gref(lookupTableGlobalName)
    ref.kill([table_name, key])

def delete_table(table_name):
    ref = iris.gref(lookupTableGlobalName)
    ref.kill([table_name])

def get_tables():
    ref = iris.gref(lookupTableGlobalName)
    key = ref.order([""])

    table_names = []

    while (key is not None):
        table_names.append(key)
        key = ref.order([key])
    
    return table_names

def list_tables():
    print("Lookup Tables:")
    for table_name in get_tables():
        print(table_name)

def load_table(table_name):
    table = {}

    ref = iris.gref(lookupTableGlobalName)
    key = ref.order([table_name, ""])

    while (key is not None):
        value = ref[table_name, key]
        table[key] = value

        key = ref.order([table_name, key])

    return table

def print_table(table_name):
    table = load_table(table_name)

    print("Lookup Table: " + table_name + '\n')
    
    print("{:<10} {:<10}".format('Key', 'Value'))
    for key, value in table.items():
        print("{:<10} {:<10}".format(key, value))

def load_xlsx(path):
    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    
    title = sheet_obj.title

    items = {}

    for i in range(1, m_row + 1):
        cell_obj_key = sheet_obj.cell(row = i, column = 1)
        cell_obj_value = sheet_obj.cell(row = i, column = 2)

        key = cell_obj_key.value
        value = cell_obj_value.value

        if (i == 1 and ("key" in str(key).lower() or "value" in str(value).lower())):
            continue

        items[key] = value
    
    return (title, items)
        
def create_lookuptable(table_name, items):

    for key, value in items.items():
        set_value(table_name, key, value)

def main():
    if (len(sys.argv) < 2):
        print("Please run this script as one of the commands below:")
        print("irispython pylotaut/pylotaut.py list")
        print("irispython pylotaut/pylotaut.py print lookup_table_name")
        print("irispython pylotaut/pylotaut.py get lookup_table_name key")

        print("irispython pylotaut/pylotaut.py set lookup_table_name key value")
        print("irispython pylotaut/pylotaut.py import path_to_xlsx")

        print("irispython pylotaut/pylotaut.py delkey lookup_table_name key")
        print("irispython pylotaut/pylotaut.py deltab lookup_table_name")

        exit(1)

    command = sys.argv[1]

    if (command == "list"):
        list_tables()

    elif (command == "print"):
        table_name = sys.argv[2]
        print_table(table_name)

    elif (command == "get"):
        table_name = sys.argv[2]
        key = sys.argv[3]
        print_value(table_name, key)

    elif (command == "set"):
        table_name = sys.argv[2]
        key = sys.argv[3]
        value = sys.argv[4]
        set_value(table_name, key, value)

    elif (command == "delkey"):
        table_name = sys.argv[2]
        key = sys.argv[3]
        delete_key(table_name, key)

    elif (command == "deltab"):
        table_name = sys.argv[2]
        delete_table(table_name)

    elif (command == "import"):
        path = sys.argv[2]
        title, items = load_xlsx("/irisdev/app/example_lookup_table.xlsx")
        create_lookuptable(title, items)

    else:
        print("Command not supported.")

if __name__ == "__main__":
    main()
